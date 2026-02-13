"""
PharmaVerif - Service de parsing des fichiers EMAC
Copyright (c) 2026 Anas BENDAIKHA
Tous droits reserves.

Fichier : backend/app/services/emac_parser.py
Parse les fichiers Excel/CSV d'EMAC pour en extraire les montants declares.

Formats supportes :
- Excel (.xlsx, .xls) via openpyxl
- CSV (.csv) via le module csv standard
- Saisie manuelle (pas de parsing, passe directement les montants)

Pattern de detection des colonnes :
  - Recherche par mots-cles dans les headers
  - Fallback sur les positions les plus courantes
"""

import csv
import io
import re
import logging
from pathlib import Path
from datetime import date, datetime
from typing import Dict, List, Optional, Any, Tuple

import openpyxl

logger = logging.getLogger(__name__)


# ========================================
# MAPPING DES COLONNES PAR MOTS-CLES
# ========================================

# Mots-cles pour detecter les colonnes dans les headers
COLUMN_KEYWORDS = {
    "reference": ["ref", "reference", "n°", "numero", "num"],
    "ca": ["ca", "chiffre", "achats", "ventes", "montant ht", "total ht"],
    "rfa": ["rfa", "remise fin", "ristourne"],
    "cop": ["cop", "objectif", "promotionnel", "condition objectif"],
    "remise_differee": ["remise diff", "diff", "remise complementaire", "compl"],
    "autres": ["autre", "divers", "escompte", "gratuit", "franco"],
    "total": ["total avantage", "total", "somme"],
    "verse": ["verse", "paye", "credite", "regle"],
    "solde": ["solde", "reste", "a percevoir", "du"],
    "periode": ["periode", "mois", "date", "trimestre"],
}


class EMACParserResult:
    """Resultat du parsing d'un fichier EMAC"""

    def __init__(self):
        self.success: bool = False
        self.message: str = ""
        self.warnings: List[str] = []

        # Montants extraits
        self.reference: Optional[str] = None
        self.ca_declare: float = 0.0
        self.rfa_declaree: float = 0.0
        self.cop_declaree: float = 0.0
        self.remises_differees_declarees: float = 0.0
        self.autres_avantages: float = 0.0
        self.total_avantages_declares: float = 0.0
        self.montant_deja_verse: float = 0.0
        self.solde_a_percevoir: float = 0.0
        self.mode_reglement: Optional[str] = None

        # Periode detectee
        self.periode_debut: Optional[date] = None
        self.periode_fin: Optional[date] = None
        self.type_periode: str = "mensuel"

        # Detail brut (lignes du fichier)
        self.detail_avantages: List[Dict[str, Any]] = []

        # Metadata
        self.format_source: str = "excel"
        self.nb_lignes_lues: int = 0

    def to_dict(self) -> Dict[str, Any]:
        """Convertir en dict pour la creation de l'EMAC"""
        return {
            "reference": self.reference,
            "ca_declare": self.ca_declare,
            "rfa_declaree": self.rfa_declaree,
            "cop_declaree": self.cop_declaree,
            "remises_differees_declarees": self.remises_differees_declarees,
            "autres_avantages": self.autres_avantages,
            "total_avantages_declares": self.total_avantages_declares,
            "montant_deja_verse": self.montant_deja_verse,
            "solde_a_percevoir": self.solde_a_percevoir,
            "mode_reglement": self.mode_reglement,
            "detail_avantages": self.detail_avantages,
            "format_source": self.format_source,
        }


class EMACParser:
    """
    Parser de fichiers EMAC (Excel / CSV)

    Detecte automatiquement les colonnes par mots-cles dans les headers,
    puis extrait les montants declares.
    """

    def parse_file(self, file_path: str) -> EMACParserResult:
        """
        Parse un fichier EMAC (Excel ou CSV).

        Args:
            file_path: Chemin vers le fichier

        Returns:
            EMACParserResult avec les montants extraits
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext in (".xlsx", ".xls"):
            return self._parse_excel(file_path)
        elif ext == ".csv":
            return self._parse_csv(file_path)
        else:
            result = EMACParserResult()
            result.message = f"Format non supporte : {ext}. Formats acceptes : .xlsx, .xls, .csv"
            return result

    def parse_bytes(self, content: bytes, filename: str) -> EMACParserResult:
        """
        Parse un fichier EMAC depuis des bytes en memoire.

        Args:
            content: Contenu brut du fichier
            filename: Nom du fichier (pour detection du format)

        Returns:
            EMACParserResult
        """
        ext = Path(filename).suffix.lower()

        if ext in (".xlsx", ".xls"):
            return self._parse_excel_bytes(content)
        elif ext == ".csv":
            return self._parse_csv_bytes(content)
        else:
            result = EMACParserResult()
            result.message = f"Format non supporte : {ext}"
            return result

    # ========================================
    # PARSING EXCEL
    # ========================================

    def _parse_excel(self, file_path: str) -> EMACParserResult:
        """Parse un fichier Excel EMAC"""
        result = EMACParserResult()
        result.format_source = "excel"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([self._clean_cell(c) for c in row])

            wb.close()
            return self._process_rows(rows, result)

        except Exception as e:
            logger.error(f"Erreur parsing Excel EMAC: {e}")
            result.message = f"Erreur lors du parsing Excel : {str(e)}"
            return result

    def _parse_excel_bytes(self, content: bytes) -> EMACParserResult:
        """Parse un fichier Excel depuis des bytes"""
        result = EMACParserResult()
        result.format_source = "excel"

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active

            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([self._clean_cell(c) for c in row])

            wb.close()
            return self._process_rows(rows, result)

        except Exception as e:
            logger.error(f"Erreur parsing Excel bytes EMAC: {e}")
            result.message = f"Erreur lors du parsing Excel : {str(e)}"
            return result

    # ========================================
    # PARSING CSV
    # ========================================

    def _parse_csv(self, file_path: str) -> EMACParserResult:
        """Parse un fichier CSV EMAC"""
        result = EMACParserResult()
        result.format_source = "csv"

        try:
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter=";")
                rows = []
                for row in reader:
                    rows.append([self._clean_cell(c) for c in row])

            return self._process_rows(rows, result)

        except UnicodeDecodeError:
            # Retry with latin-1
            try:
                with open(file_path, "r", encoding="latin-1") as f:
                    reader = csv.reader(f, delimiter=";")
                    rows = []
                    for row in reader:
                        rows.append([self._clean_cell(c) for c in row])

                return self._process_rows(rows, result)
            except Exception as e:
                result.message = f"Erreur encodage CSV : {str(e)}"
                return result

        except Exception as e:
            logger.error(f"Erreur parsing CSV EMAC: {e}")
            result.message = f"Erreur lors du parsing CSV : {str(e)}"
            return result

    def _parse_csv_bytes(self, content: bytes) -> EMACParserResult:
        """Parse un fichier CSV depuis des bytes"""
        result = EMACParserResult()
        result.format_source = "csv"

        try:
            # Essayer utf-8 puis latin-1
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("latin-1")

            reader = csv.reader(io.StringIO(text), delimiter=";")
            rows = []
            for row in reader:
                rows.append([self._clean_cell(c) for c in row])

            return self._process_rows(rows, result)

        except Exception as e:
            logger.error(f"Erreur parsing CSV bytes EMAC: {e}")
            result.message = f"Erreur lors du parsing CSV : {str(e)}"
            return result

    # ========================================
    # TRAITEMENT DES LIGNES
    # ========================================

    def _process_rows(self, rows: List[List[Any]], result: EMACParserResult) -> EMACParserResult:
        """
        Traite les lignes extraites du fichier.

        Strategie :
        1. Chercher la ligne header (mots-cles)
        2. Si trouvee, mapper les colonnes et extraire les lignes de donnees
        3. Sinon, utiliser une strategie de recherche par paires cle-valeur
        """
        if not rows or len(rows) < 2:
            result.message = "Fichier vide ou trop peu de donnees"
            return result

        result.nb_lignes_lues = len(rows)

        # Strategie 1 : Chercher des paires cle-valeur (format "label: valeur" courant dans les EMAC)
        kv_result = self._extract_key_value_pairs(rows)
        if kv_result:
            self._apply_kv_results(kv_result, result)
            if result.ca_declare > 0 or result.total_avantages_declares > 0:
                result.success = True
                result.message = f"EMAC parse avec succes (format cle-valeur, {len(kv_result)} champs detectes)"
                self._auto_calculate(result)
                return result

        # Strategie 2 : Chercher header + lignes tabulaires
        header_idx, column_map = self._find_header(rows)
        if header_idx is not None and column_map:
            self._extract_tabular_data(rows, header_idx, column_map, result)
            if result.ca_declare > 0 or result.total_avantages_declares > 0:
                result.success = True
                result.message = f"EMAC parse avec succes (format tabulaire, {len(column_map)} colonnes detectees)"
                self._auto_calculate(result)
                return result

        # Strategie 3 : Recherche brute de montants
        self._extract_brute(rows, result)
        if result.ca_declare > 0 or result.total_avantages_declares > 0:
            result.success = True
            result.message = "EMAC parse avec extraction brute des montants"
            result.warnings.append("Extraction brute : verifier manuellement les montants")
            self._auto_calculate(result)
            return result

        result.message = "Impossible d'extraire les donnees EMAC du fichier"
        result.warnings.append("Aucun montant detecte. Saisie manuelle recommandee.")
        return result

    def _find_header(self, rows: List[List[Any]]) -> Tuple[Optional[int], Dict[str, int]]:
        """
        Cherche la ligne header et mappe les colonnes.

        Returns:
            (index_header, {nom_champ: index_colonne})
        """
        for i, row in enumerate(rows[:10]):  # Ne chercher que dans les 10 premieres lignes
            column_map: Dict[str, int] = {}
            for j, cell in enumerate(row):
                if not cell:
                    continue
                cell_lower = str(cell).lower().strip()

                for field, keywords in COLUMN_KEYWORDS.items():
                    if any(kw in cell_lower for kw in keywords):
                        if field not in column_map:  # Premier match gagne
                            column_map[field] = j

            # Si on a detecte au moins 2 colonnes significatives, c'est le header
            significant = {"ca", "rfa", "cop", "total", "remise_differee"}
            if len(column_map.keys() & significant) >= 2:
                return i, column_map

        return None, {}

    def _extract_tabular_data(
        self,
        rows: List[List[Any]],
        header_idx: int,
        column_map: Dict[str, int],
        result: EMACParserResult,
    ) -> None:
        """Extrait les donnees des lignes tabulaires apres le header"""
        for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            line_data: Dict[str, Any] = {}
            for field, col_idx in column_map.items():
                if col_idx < len(row):
                    line_data[field] = row[col_idx]

            # Accumuler les montants
            if "ca" in line_data:
                val = self._parse_number(line_data["ca"])
                if val is not None:
                    result.ca_declare += val
            if "rfa" in line_data:
                val = self._parse_number(line_data["rfa"])
                if val is not None:
                    result.rfa_declaree += val
            if "cop" in line_data:
                val = self._parse_number(line_data["cop"])
                if val is not None:
                    result.cop_declaree += val
            if "remise_differee" in line_data:
                val = self._parse_number(line_data["remise_differee"])
                if val is not None:
                    result.remises_differees_declarees += val
            if "autres" in line_data:
                val = self._parse_number(line_data["autres"])
                if val is not None:
                    result.autres_avantages += val
            if "total" in line_data:
                val = self._parse_number(line_data["total"])
                if val is not None:
                    result.total_avantages_declares = val  # Prendre la derniere valeur
            if "verse" in line_data:
                val = self._parse_number(line_data["verse"])
                if val is not None:
                    result.montant_deja_verse += val
            if "solde" in line_data:
                val = self._parse_number(line_data["solde"])
                if val is not None:
                    result.solde_a_percevoir = val

            result.detail_avantages.append(line_data)

    def _extract_key_value_pairs(self, rows: List[List[Any]]) -> Dict[str, Any]:
        """
        Extrait des paires cle-valeur (format courant des EMAC).
        Ex: "CA HT" | 125000.00
             "RFA"   | 3750.00
        """
        kv: Dict[str, Any] = {}

        for row in rows:
            if not row or len(row) < 2:
                continue

            # Chercher pattern : [label, valeur]
            label = str(row[0]).strip().lower() if row[0] else ""
            if not label:
                continue

            # Chercher la premiere valeur numerique dans la ligne
            value = None
            for cell in row[1:]:
                v = self._parse_number(cell)
                if v is not None:
                    value = v
                    break

            if value is None:
                continue

            # Mapper le label aux champs EMAC
            for field, keywords in COLUMN_KEYWORDS.items():
                if any(kw in label for kw in keywords):
                    kv[field] = value
                    break

        return kv

    def _apply_kv_results(self, kv: Dict[str, Any], result: EMACParserResult) -> None:
        """Applique les paires cle-valeur au resultat"""
        if "ca" in kv:
            result.ca_declare = float(kv["ca"])
        if "rfa" in kv:
            result.rfa_declaree = float(kv["rfa"])
        if "cop" in kv:
            result.cop_declaree = float(kv["cop"])
        if "remise_differee" in kv:
            result.remises_differees_declarees = float(kv["remise_differee"])
        if "autres" in kv:
            result.autres_avantages = float(kv["autres"])
        if "total" in kv:
            result.total_avantages_declares = float(kv["total"])
        if "verse" in kv:
            result.montant_deja_verse = float(kv["verse"])
        if "solde" in kv:
            result.solde_a_percevoir = float(kv["solde"])
        if "reference" in kv:
            result.reference = str(kv["reference"])

    def _extract_brute(self, rows: List[List[Any]], result: EMACParserResult) -> None:
        """
        Extraction brute : cherche les plus gros montants dans le fichier.
        Utilise en dernier recours quand aucun pattern n'est detecte.
        """
        all_numbers: List[Tuple[float, str]] = []

        for row in rows:
            for cell in row:
                val = self._parse_number(cell)
                if val is not None and val > 0:
                    label = ""
                    # Chercher un label dans la meme ligne
                    for c in row:
                        if c and not self._parse_number(c) and str(c).strip():
                            label = str(c).strip()
                            break
                    all_numbers.append((val, label))

        # Trier par montant decroissant
        all_numbers.sort(key=lambda x: x[0], reverse=True)

        if all_numbers:
            # Le plus gros montant est probablement le CA
            result.ca_declare = all_numbers[0][0]
            result.detail_avantages = [
                {"montant": n, "label": l} for n, l in all_numbers[:20]
            ]

            # Chercher un total avantages (montant significatif plus petit que le CA)
            for val, label in all_numbers[1:]:
                label_lower = label.lower()
                if "total" in label_lower or "avantage" in label_lower:
                    result.total_avantages_declares = val
                    break
                elif "rfa" in label_lower or "ristourne" in label_lower:
                    result.rfa_declaree = val

    # ========================================
    # HELPERS
    # ========================================

    def _auto_calculate(self, result: EMACParserResult) -> None:
        """Auto-calcule les champs manquants"""
        # Total avantages = somme si pas fourni
        if result.total_avantages_declares == 0:
            result.total_avantages_declares = (
                result.rfa_declaree
                + result.cop_declaree
                + result.remises_differees_declarees
                + result.autres_avantages
            )

        # Solde = total - verse si pas fourni
        if result.solde_a_percevoir == 0 and result.total_avantages_declares > 0:
            result.solde_a_percevoir = max(
                0.0,
                result.total_avantages_declares - result.montant_deja_verse,
            )

    @staticmethod
    def _clean_cell(cell: Any) -> Any:
        """Nettoie une cellule"""
        if cell is None:
            return None
        if isinstance(cell, str):
            cell = cell.strip()
            if cell == "":
                return None
        return cell

    @staticmethod
    def _parse_number(value: Any) -> Optional[float]:
        """
        Parse une valeur en nombre flottant.
        Gere les formats : 1234.56, 1 234,56, 1234,56, "1234.56 EUR"
        """
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if not isinstance(value, str):
            return None

        # Nettoyer
        s = value.strip()
        # Retirer les devises et suffixes
        s = re.sub(r'[€$£EUR\s]', '', s, flags=re.IGNORECASE)
        # Retirer les espaces (separateurs de milliers)
        s = s.replace(' ', '').replace('\u00a0', '')

        if not s:
            return None

        # Format francais : 1234,56 ou 1.234,56
        if ',' in s and '.' in s:
            # 1.234,56 -> 1234.56
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            # 1234,56 -> 1234.56
            s = s.replace(',', '.')

        try:
            return float(s)
        except ValueError:
            return None
